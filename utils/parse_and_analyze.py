"""合并解析+分析脚本：直接从 raw 文件生成 _analysis.json，跳过中间 Markdown。

对照目的：与标准两阶段方案（parser → analyze）对比，评估：
  1. 中间 Markdown 是否对变更检测质量有增益
  2. 合并流程的速度差异
  3. 两种方案的精度取舍

用法：
    python utils/parse_and_analyze.py                         # 处理所有文件
    python utils/parse_and_analyze.py --source sse --limit 5  # SSE 前5个
    python utils/parse_and_analyze.py --dry-run               # 试运行
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.FileHandler(Path("log/parse_and_analyze.log"), encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("parse_and_analyze")

RAW_DIR = Path("knowledge/raw")
OUT_DIR = Path("knowledge/articles/parseranalyzer")

DOC_TYPES = {
    "规格说明书": "interface_spec", "接口规范": "interface_spec",
    "数据接口": "interface_spec", "接口规格": "interface_spec",
    "技术实施指南": "guide", "实施指南": "guide", "技术指南": "guide",
    "通知": "technical_notice", "测试方案": "test_doc", "测试": "test_doc",
    "软件": "software", "下载": "software", "错误代码表": "software",
    "杂志": "magazine", "业务规则": "business_rule",
}

CAT_MAP = {
    "技术通知": "technical_notice", "服务指引": "guide",
    "技术接口": "interface_spec", "技术指南": "guide",
    "软件下载": "software", "测试文档": "test_doc",
    "技术杂志": "magazine", "历史资料": "guide",
    "技术公告": "technical_notice", "交易系统介绍": "guide",
    "数据接口": "interface_spec", "业务规则": "business_rule",
}

CHANGE_TYPE_KW = {
    "接口字段": "接口字段变更", "字段长度": "接口字段变更",
    "字段类型": "接口字段变更", "必填": "接口字段变更",
    "枚举": "接口字段变更", "代码": "接口字段变更",
    "流程": "业务流程变更", "步骤": "业务流程变更",
    "规则": "规则条款变更", "条款": "规则条款变更",
    "架构": "技术架构变更", "通信": "技术架构变更",
    "协议": "技术架构变更", "FTP": "技术架构变更",
    "版本": "版本升级", "V1": "版本升级",
    "废止": "废止", "停止使用": "废止", "不再支持": "废止", "下线": "废止",
}

SEVERITY_KW = {
    "critical": ["安全", "风险", "数据丢失"],
    "major": ["新增", "删除", "废止", "修改", "变更", "替换", "迁移"],
    "minor": ["优化", "说明", "补充", "扩容"],
    "cosmetic": ["格式", "排版", "文案", "更正"],
}

_CRAWL_META = None


def _load_crawl_meta():
    global _CRAWL_META
    if _CRAWL_META is None:
        p = RAW_DIR / "crawl_metadata.json"
        if p.exists():
            _CRAWL_META = json.loads(p.read_text("utf-8"))
        else:
            _CRAWL_META = []
    return _CRAWL_META


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _infer_doc_type(title: str, category: str) -> str:
    for kw, dt in DOC_TYPES.items():
        if kw in title:
            return dt
    return CAT_MAP.get(category, "technical_notice")


def _infer_version(title: str) -> str | None:
    for p in [r"V(\d+[\.\d]*)", r"v(\d+[\.\d]*)", r"Ver(\d+[\.\d]*)",
              r"(\d+[\.\d]*)版", r"版本(\d+[\.\d]*)"]:
        m = re.search(p, title)
        if m:
            return m.group(1)
    return None


def _extract_date_from_filename(fname: str) -> str | None:
    m = re.match(r"(\d{4})(\d{2})(\d{2})_", fname)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.search(r"(\d{4})(\d{2})(\d{2})", fname)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def _color_to_css(color: int) -> str | None:
    c = color & 0xFFFFFF
    r, g, b = (c >> 16) & 0xFF, (c >> 8) & 0xFF, c & 0xFF
    if r > 200 and g < 80 and b < 80:
        return "red"
    if b > 200 and r < 80 and g < 80:
        return "blue"
    if r < 60 and g < 60 and b < 60:
        return None
    return None


def _classify_change_type(text: str) -> str:
    for kw, ct in CHANGE_TYPE_KW.items():
        if kw in text:
            return ct
    return "规则条款变更"


def _classify_severity(text: str) -> str:
    for sev, kws in SEVERITY_KW.items():
        for kw in kws:
            if kw in text:
                return sev
    return "minor"


def _is_valid_change_text(text: str) -> bool:
    if len(text) < 4:
        return False
    if all(c in "，。、；：？！""''（）【】《》—…·\n\r\t " for c in text):
        return False
    return True


def _is_page_artifact(text: str) -> bool:
    t = text.strip()
    if re.match(r"^\d{1,3}$", t):
        return True
    if re.match(r"^第\s*\d+\s*页", t):
        return True
    if re.match(r"^(上海证券交易所|深圳证券交易所)", t) and len(t) < 30:
        return True
    return False


def extract_pdf_text_and_changes(path: Path) -> tuple[str, list[dict]]:
    import fitz
    doc = None
    try:
        doc = fitz.open(path)
        raw_text_parts: list[str] = []
        changes: list[dict] = []
        seen = set()
        for page in doc:
            blocks = page.get_text("dict", sort=True).get("blocks", [])
            for block in blocks:
                if block.get("type") != 0:
                    continue
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        text = span.get("text", "").strip()
                        if not text:
                            continue
                        raw_text_parts.append(text)
                        color = span.get("color", 0)
                        css_color = _color_to_css(color)
                        if not css_color:
                            continue
                        if not _is_valid_change_text(text):
                            continue
                        if _is_page_artifact(text):
                            continue
                        dedup = f"{css_color}|{text[:80]}"
                        if dedup in seen:
                            continue
                        seen.add(dedup)
                        ct = _classify_change_type(text)
                        sev = _classify_severity(text)
                        if css_color == "blue" and sev == "major":
                            sev = "minor"
                        changes.append({
                            "type": ct, "summary": text[:200],
                            "detail": text[:400], "severity": sev,
                            "source": "color_annotation",
                        })
        raw_text = "\n".join(raw_text_parts[-200:])
        return raw_text, changes
    finally:
        if doc:
            doc.close()


def extract_docx_text_and_changes(path: Path) -> tuple[str, list[dict]]:
    import docx as docx_lib
    d = docx_lib.Document(path)
    raw_text_parts: list[str] = []
    changes: list[dict] = []
    for para in d.paragraphs:
        text = para.text.strip()
        if text:
            raw_text_parts.append(text)
    # docx color detection is limited; fallback to keyword
    for line in raw_text_parts:
        for kw in ["新增", "修改", "删除", "废止", "停止使用"]:
            if kw in line:
                changes.append({
                    "type": kw, "summary": line[:200],
                    "detail": line[:400], "severity": "major",
                    "source": "keyword_detect",
                })
                break
    return "\n".join(raw_text_parts[-200:]), changes


def extract_xlsx_text(path: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    parts = []
    for name in wb.sheetnames:
        parts.append(f"[Sheet: {name}]")
        rows = list(wb[name].iter_rows(values_only=True))
        for row in rows[:20]:
            parts.append(" | ".join(str(c or "") for c in row))
    wb.close()
    return "\n".join(parts)


def extract_html_text(path: Path) -> str:
    from bs4 import BeautifulSoup
    html = path.read_text("utf-8", errors="replace")
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "nav", "footer"]):
        tag.decompose()
    return soup.get_text(separator="\n", strip=True)


def find_source_url(fname: str) -> str | None:
    for item in _load_crawl_meta():
        lp = item.get("local_path") or ""
        u = item.get("url") or ""
        if fname in lp or fname in u:
            return u
        title = item.get("title") or ""
        if title and title[:20] in fname:
            return u
    return None


def collect_raw_files(source: str | None = None, limit: int | None = None) -> list[tuple[str, str, str | None, Path]]:
    known_exts = {".pdf", ".docx", ".xlsx", ".xls", ".html", ".shtml"}
    files: list[tuple[str, str, str | None, Path]] = []
    for raw_dir in RAW_DIR.iterdir():
        if not raw_dir.is_dir():
            continue
        src = raw_dir.name
        if source and src != source:
            continue
        for cat_dir in raw_dir.iterdir():
            if not cat_dir.is_dir():
                continue
            cat = cat_dir.name
            sub = None
            if src == "chinaclear":
                sub, cat = cat, "业务规则"
            for f in sorted(cat_dir.iterdir()):
                if f.is_file() and f.suffix.lower() in known_exts:
                    files.append((src, cat, sub, f))
    if limit:
        files = files[:limit]
    return files


def seq_for_date(date_str: str, existing: set) -> int:
    prefix = date_str.replace("-", "")
    count = sum(1 for e in existing if prefix in e)
    return count + 1


def process_file(source: str, category: str, sub_category: str | None, path: Path, existing_ids: set) -> dict | None:
    fname = path.name
    try:
        ext = path.suffix.lower()
        if ext == ".pdf":
            raw_text, changes = extract_pdf_text_and_changes(path)
        elif ext == ".docx":
            raw_text, changes = extract_docx_text_and_changes(path)
        elif ext in (".xlsx", ".xls"):
            raw_text = extract_xlsx_text(path)
            changes = []
        elif ext in (".html", ".shtml"):
            raw_text = extract_html_text(path)
            changes = []
        else:
            return None
    except Exception as e:
        logger.error("Extract failed [%s]: %s", fname, e)
        return None

    title = re.sub(r"^\d{8}_", "", path.stem)
    doc_type = _infer_doc_type(title, category)
    version = _infer_version(title)
    public_date = _extract_date_from_filename(fname)
    file_hash = _sha256(path)
    source_url = find_source_url(fname)

    # Tags
    tags = [source, doc_type]
    for kw in [r"IS\d{3}", r"STEP", r"BINARY", r"ETF", r"UniTrans", r"EzOES",
               r"期权", r"债券", r"科创板", r"REITs", r"行情网关", r"交易网关"]:
        if re.search(kw, title):
            tags.append(re.search(kw, title).group())
    if changes:
        tags.append("has_changes")
    tags = list(dict.fromkeys(tags))[:10]

    # Status
    status = "active"
    dep_date = None
    sup_by = None
    for c in changes:
        if c["type"] == "废止":
            status = "deprecated"
            m = re.search(r"(\d{4}[-.]\d{1,2}[-.]\d{1,2})", c["detail"])
            if m:
                dep_date = m.group(1)
            sm = re.search(r"(IS\d{3})", c["detail"])
            if sm:
                sup_by = sm.group(1)
            break

    # doc_id
    short_type = {"technical_notice": "tech", "interface_spec": "iface",
                  "business_rule": "rule", "guide": "guide",
                  "software": "soft", "test_doc": "test", "magazine": "mag"}.get(doc_type, "tech")
    date_part = (public_date or "00000000").replace("-", "")
    seq = seq_for_date(date_part, existing_ids)
    doc_id = f"{source}-{short_type}-{date_part}-{seq:03d}"
    existing_ids.add(doc_id)

    # Summary
    if not changes:
        summary = f"初始版本，无历史变更。类型：{doc_type}，版本：{version or '无'}。"
    else:
        parts = [f"本次涉及{len(changes)}项变更。"]
        for ct in set(c["type"] for c in changes):
            parts.append(f"{ct}: {sum(1 for c in changes if c['type']==ct)}项。")
        majors = sum(1 for c in changes if c["severity"] == "major")
        if majors:
            parts.append(f"其中重大变更{majors}项。")
        summary = " ".join(parts)[:200]

    confidence = 0.92 if changes else 0.72

    return {
        "doc_id": doc_id,
        "title": title,
        "source": source,
        "source_url": source_url,
        "analysis_date": datetime.now(timezone.utc).isoformat(),
        "metadata": {
            "file_hash": f"sha256:{file_hash}",
            "file_format": ext.lstrip("."),
            "page_count": 0,
            "doc_type": doc_type,
            "version": version,
            "previous_version": None,
            "public_date": public_date,
            "effective_date": None,
            "parse_status": "success",
        },
        "status": status,
        "version": version,
        "previous_version": None,
        "changes": changes,
        "tags": tags,
        "related_ids": [],
        "deprecated_date": dep_date,
        "superseded_by": sup_by,
        "summary": summary,
        "confidence": round(confidence, 2),
        "raw_text": raw_text[:5000],
    }


def main():
    parser = argparse.ArgumentParser(description="合并解析+分析（跳过中间Markdown）")
    parser.add_argument("--source", help="限定数据源: sse / szse / chinaclear")
    parser.add_argument("--limit", type=int, help="最大处理文件数")
    parser.add_argument("--dry-run", action="store_true", help="试运行，不写入")
    parser.add_argument("--request-delay", type=float, default=0.2)
    args = parser.parse_args()

    files = collect_raw_files(args.source, args.limit)
    if args.dry_run:
        logger.info("DRY RUN: %d files would be processed", len(files))
        for src, cat, sub, fp in files:
            logger.info("  [%s] [%s] %s", src, cat, fp.name)
        return

    existing_ids = set()
    ok = fail = 0
    for src, cat, sub, fp in files:
        logger.info("Processing [%s] [%s] %s", src, cat, fp.name)
        result = process_file(src, cat, sub, fp, existing_ids)
        if result is None:
            fail += 1
            continue
        out_dir = OUT_DIR / src / cat
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{result['doc_id']}_analysis.json"
        out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        has_c = "✓" if result["changes"] else " "
        logger.info("  [%s] %s", has_c, result["doc_id"])
        ok += 1
        time.sleep(args.request_delay)

    logger.info("Done: %d success, %d failed", ok, fail)


if __name__ == "__main__":
    main()
